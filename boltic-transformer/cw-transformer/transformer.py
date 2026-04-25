"""
Cottonworld Logic ERP → Fynd Platform Template Transformer

Converts Logic ERP Item Master export (.xlsx) to a Fynd Commerce Platform
bulk-upload template. Uses only openpyxl + stdlib (no pandas/numpy) to keep
build times fast on Boltic Serverless.

Returns (BytesIO, warnings, None) so the handler can surface data-quality issues.
"""

from __future__ import annotations

import csv
import json
import math
import re
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Reference data (loaded once from /data)
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).parent / "data"


def _load_json(name: str) -> dict:
    with open(DATA_DIR / name, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_hsn_lookup() -> dict[tuple[str, str], str]:
    lookup: dict[tuple[str, str], str] = {}
    with open(DATA_DIR / "hsn_lookup.csv", "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            section = row["section"].strip().upper()
            dept = row["department"].strip().upper()
            hs_code = row["hs_code"].strip()
            lookup[(section, dept)] = hs_code
    return lookup


SLEEVE_MAP = _load_json("sleeve_map.json")
COLLAR_MAP = _load_json("collar_map.json")
MATERIAL_MAP = _load_json("material_map.json")
SECTION_GENDER = _load_json("section_gender.json")
DEPARTMENT_DISPLAY = _load_json("department_display.json")
STATIC = _load_json("static_values.json")
HSN_LOOKUP = _load_hsn_lookup()


# ---------------------------------------------------------------------------
# Fynd output schema (matches Supplementary template CW - Sample.xlsx)
# ---------------------------------------------------------------------------

FYND_FIXED_COLUMNS = [
    "Name", "Slug", "Item Code", "Brand", "Category", "Description",
    "Short Description", "Tax Rule Name", "HS Code", "Country of Origin",
    "Media", "Multi Size", "Gtin Type", "Gtin Value", "Seller Identifier",
    "Meta", "Size Meta", "Size", "Actual Price", "Selling Price",
    "Currency", "Length (cm)", "Width (cm)", "Height (cm)",
    "Product Dead Weight (gram)", "Size Guide", "Available", "Highlights",
    "Unlisted Product", "Variant Type", "Variant Group ID", "Variant Media",
    "Trader Type", "Trader Name", "Trader Address", "Track Inventory",
    "Teaser Tag Name", "No of Boxes", "Manufacturing Time",
    "Manufacturing Time Unit", "Return Time Limit", "Return Time Unit",
    "Product Publishing Date", "Tags", "Net Quantity Value",
    "Net Quantity Unit", "Product Bundle", "Colour", "Material",
    "Package Contents", "Quantity Factor", "Priority",
]

FYND_CUSTOM_ATTRS = [f"Custom Attribute {i}" for i in range(1, 51)]
FYND_COLUMNS = FYND_FIXED_COLUMNS + FYND_CUSTOM_ATTRS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return False


def clean_val(val: Any) -> str:
    """Return empty string for None/NaN/'(NIL)'. Convert whole-number floats to int."""
    if _is_empty(val):
        return ""
    if isinstance(val, float) and val == int(val):
        val = int(val)
    s = str(val).strip()
    if s.upper() == "(NIL)" or s.lower() == "nan":
        return ""
    return s


def strip_percentages(composition: str) -> str:
    """'49% COTTON 46% BAMBOO 5% ELASTANE' -> 'Cotton Bamboo Elastane'."""
    if not composition:
        return ""
    text = re.sub(r"\d+\s*%\s*", "", composition)
    text = re.sub(r"\s+", " ", text).strip(" ,/-")
    return text.title()


def extract_primary_material(fabric_sub_type_raw: str) -> str:
    val = clean_val(fabric_sub_type_raw)
    if not val:
        return ""
    upper = val.upper()
    if upper in MATERIAL_MAP:
        return MATERIAL_MAP[upper]
    for key, mapped in MATERIAL_MAP.items():
        if key in upper:
            return mapped
    first = val.split("/")[0].strip()
    return first.title()


def map_sleeve_type(sleeve_raw: str, warnings: list[str]) -> str:
    val = clean_val(sleeve_raw).upper()
    if not val:
        return ""
    if val in SLEEVE_MAP:
        return SLEEVE_MAP[val]
    for key, mapped in SLEEVE_MAP.items():
        if key and key in val:
            return mapped
    warnings.append(f"Unknown sleeve type '{sleeve_raw}' — passed through as-is.")
    return val.title()


def map_collar_style(collar_raw: str, warnings: list[str]) -> str:
    val = clean_val(collar_raw).upper()
    if not val:
        return ""
    if val in COLLAR_MAP:
        return COLLAR_MAP[val]
    warnings.append(f"Unknown neck/collar '{collar_raw}' — passed through as-is.")
    return val.title()


def gender_info(section: str) -> dict:
    key = section.upper().strip()
    if key in SECTION_GENDER:
        return SECTION_GENDER[key]
    return {"gender": section.title(), "possessive": f"{section.title()}'s", "prefix": "X"}


def department_display(department: str) -> str:
    key = department.upper().strip()
    if key in DEPARTMENT_DISPLAY:
        return DEPARTMENT_DISPLAY[key]
    return department.title()


def build_product_name(section: str, composition1: str, fit: str,
                       department: str, color: str) -> str:
    info = gender_info(section)
    parts: list[str] = [info["possessive"]]
    comp = strip_percentages(clean_val(composition1))
    if comp:
        parts.append(comp)
    fit_clean = clean_val(fit).title()
    if fit_clean:
        parts.append(fit_clean)
    dept_disp = department_display(clean_val(department))
    if dept_disp:
        parts.append(dept_disp)
    color_clean = clean_val(color).title()
    if color_clean:
        parts.append(color_clean)
    return " ".join(parts)


def build_item_code(section: str, department: str, style_no: str,
                    fabric_no: str, color: str) -> str:
    info = gender_info(section)
    prefix = info["prefix"]
    dept = clean_val(department).upper().replace(" ", "-")
    style = clean_val(style_no)
    fabric = clean_val(fabric_no)
    col = clean_val(color).upper().replace(" ", "-")
    return f"{prefix}-{dept}-{style}-{fabric}-{col}"


def lookup_hs_code(section: str, department: str, warnings: list[str]) -> str:
    sec = clean_val(section).upper()
    dept = clean_val(department).upper()
    if not sec or not dept:
        return ""
    hs = HSN_LOOKUP.get((sec, dept))
    if hs is None:
        warnings.append(
            f"No HSN mapping for Section='{sec}', Department='{dept}'. "
            f"HS Code left blank — please add to data/hsn_lookup.csv."
        )
        return ""
    return hs


def format_size(size_raw: str) -> str:
    val = clean_val(size_raw).upper()
    size_shortforms = {
        "SMALL": "S", "MEDIUM": "M", "LARGE": "L",
        "XLARGE": "XL", "XXLARGE": "XXL", "XSMALL": "XS",
        "X-LARGE": "XL", "X-SMALL": "XS", "XX-LARGE": "XXL",
    }
    return size_shortforms.get(val, val)


def clean_barcode(raw: Any) -> str:
    val = clean_val(raw)
    if not val:
        return ""
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return val


# ---------------------------------------------------------------------------
# Column discovery (Logic column names vary in case/spacing)
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def find_col(headers: list[str], *candidates: str) -> int | None:
    """Return index of first matching header; None if not found."""
    norm_map = {_normalize(h): i for i, h in enumerate(headers)}
    for cand in candidates:
        n = _normalize(cand)
        if n in norm_map:
            return norm_map[n]
    return None


# ---------------------------------------------------------------------------
# Main transform
# ---------------------------------------------------------------------------

def transform(input_file) -> tuple[BytesIO, list[str], None]:
    """
    Read Logic ERP xlsx, transform to Fynd Platform template.
    Returns (BytesIO of output xlsx, list of warning strings, None).
    """
    warnings: list[str] = []

    # Load workbook with openpyxl
    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    # Find header row containing OEM_BARCODE
    header_idx = None
    for i, row in enumerate(all_rows):
        for val in row:
            if str(val).strip().upper() == "OEM_BARCODE":
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is None:
        raise ValueError("Could not find 'OEM_BARCODE' header row in input file.")

    headers = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    # Resolve column indices
    col_oem       = find_col(headers, "OEM_BARCODE")
    col_section   = find_col(headers, "SECTION")
    col_dept      = find_col(headers, "DEPARTMENT")
    col_style_no  = find_col(headers, "STYLE NO", "STYLE_NO", "STYLENO")
    col_fabric_no = find_col(headers, "FABRIC NO.", "FABRIC_NO", "FABRICNO")
    col_color     = find_col(headers, "COLOR")
    col_size      = find_col(headers, "PACK / SIZE", "PACK_SIZE", "PACKSIZE", "PACK/SIZE")
    col_fab_main  = find_col(headers, "FABRIC MAIN DESC", "FABRIC_MAIN_DESC")
    col_fab_sub   = find_col(headers, "FABRIC SUB DESC", "FABRIC_SUB_DESC")
    col_fab_type  = find_col(headers, "FABRIC TYPE", "FABRIC_TYPE")
    col_fab_subtype = find_col(headers, "FABRIC SUB TYPE", "FABRIC_SUB_TYPE")
    col_hl        = find_col(headers, "HL")
    col_sleeve    = find_col(headers, "SLEEVE TYPE", "SLEEVE_TYPE")
    col_fit       = find_col(headers, "FIT")
    col_occasion  = find_col(headers, "OCCASION")
    col_pockets   = find_col(headers, "POCKETS")
    col_neck      = find_col(headers, "NECK-COLLAR", "NECK_COLLAR", "NECKCOLLAR")
    col_length    = find_col(headers, "LENGTH")
    col_waist     = find_col(headers, "WAIST")
    col_closure   = find_col(headers, "CLOSURE")
    col_leg       = find_col(headers, "LEG")
    col_front     = find_col(headers, "FRONT")
    col_comp1     = find_col(headers, "COMPOSITION1")
    col_packed_date = find_col(headers, "PACKED DATE", "PACKED_DATE")
    col_cs        = find_col(headers, "CS")
    col_rate      = find_col(headers, "RATE")
    col_order_no  = find_col(headers, "ORDER NO", "ORDER_NO")
    col_mrp       = find_col(headers, "MRP")

    # Validate required columns
    required = {
        "OEM_BARCODE": col_oem, "SECTION": col_section, "DEPARTMENT": col_dept,
        "STYLE NO": col_style_no, "FABRIC NO.": col_fabric_no,
        "COLOR": col_color, "PACK / SIZE": col_size, "MRP": col_mrp,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"Input file missing required columns: {', '.join(missing)}")

    def get(row: list, idx: int | None) -> Any:
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Filter rows with OEM_BARCODE present
    valid_rows = [r for r in data_rows if not _is_empty(get(r, col_oem))]

    # Group by style_no + fabric_no + color (preserving order)
    groups: dict[str, list] = {}
    group_order: list[str] = []
    for row in valid_rows:
        key = (
            clean_val(get(row, col_style_no)) + "|" +
            clean_val(get(row, col_fabric_no)) + "|" +
            clean_val(get(row, col_color))
        )
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(row)

    output_rows: list[dict] = []

    for key in group_order:
        group = groups[key]
        first = group[0]

        section      = clean_val(get(first, col_section))
        department   = clean_val(get(first, col_dept))
        style_no     = clean_val(get(first, col_style_no))
        fabric_no    = clean_val(get(first, col_fabric_no))
        color        = clean_val(get(first, col_color))
        fit          = clean_val(get(first, col_fit))
        occasion     = clean_val(get(first, col_occasion))
        neck_collar  = clean_val(get(first, col_neck))
        sleeve_type  = clean_val(get(first, col_sleeve))
        fabric_main  = clean_val(get(first, col_fab_main))
        fabric_sub   = clean_val(get(first, col_fab_sub))
        fabric_type  = clean_val(get(first, col_fab_type))
        fabric_sub_type = clean_val(get(first, col_fab_subtype))
        hl           = clean_val(get(first, col_hl))
        pockets      = clean_val(get(first, col_pockets))
        length_val   = clean_val(get(first, col_length))
        waist_val    = clean_val(get(first, col_waist))
        closure      = clean_val(get(first, col_closure))
        leg          = clean_val(get(first, col_leg))
        front        = clean_val(get(first, col_front))
        composition1 = clean_val(get(first, col_comp1))
        packed_date  = clean_val(get(first, col_packed_date))
        cs           = clean_val(get(first, col_cs))
        rate         = clean_val(get(first, col_rate))
        order_no     = clean_val(get(first, col_order_no))

        product_name = build_product_name(section, composition1, fit, department, color)
        item_code    = build_item_code(section, department, style_no, fabric_no, color)
        hs_code      = lookup_hs_code(section, department, warnings)
        info         = gender_info(section)
        gender       = info["gender"]
        dept_disp    = department_display(department)
        primary_material_fabric = extract_primary_material(fabric_sub_type)
        material_display = strip_percentages(composition1)
        collar_mapped = map_collar_style(neck_collar, warnings)
        sleeve_mapped = map_sleeve_type(sleeve_type, warnings)
        fit_display  = fit.title() if fit else ""

        for i, row in enumerate(group):
            is_first = (i == 0)
            oem_barcode = clean_barcode(get(row, col_oem))
            size = format_size(clean_val(get(row, col_size)))

            mrp_raw = get(row, col_mrp)
            if _is_empty(mrp_raw):
                mrp_val: Any = ""
            else:
                try:
                    mrp_val = int(float(str(mrp_raw)))
                except (ValueError, TypeError):
                    mrp_val = ""

            out = {col: "" for col in FYND_COLUMNS}

            # Variant-level (every row)
            out["Item Code"]                  = item_code
            out["Brand"]                      = STATIC["brand"]
            out["Gtin Type"]                  = STATIC["gtin_type"]
            out["Gtin Value"]                 = oem_barcode
            out["Seller Identifier"]          = oem_barcode
            out["Size"]                       = size
            out["Actual Price"]               = mrp_val
            out["Selling Price"]              = mrp_val
            out["Currency"]                   = STATIC["currency"]
            out["Length (cm)"]                = STATIC["length_cm"]
            out["Width (cm)"]                 = STATIC["width_cm"]
            out["Height (cm)"]                = STATIC["height_cm"]
            out["Product Dead Weight (gram)"] = STATIC["weight_gram"]

            # Product-level (first row of each group only)
            if is_first:
                out["Name"]              = product_name
                out["Category"]          = STATIC["category"]
                out["Tax Rule Name"]     = STATIC["tax_rule"]
                out["HS Code"]           = hs_code
                out["Country of Origin"] = STATIC["country_of_origin"]
                out["Trader Type"]       = STATIC["trader_type"]
                out["Trader Name"]       = STATIC["trader_name"]
                out["Trader Address"]    = STATIC["trader_address"]
                out["Return Time Limit"] = STATIC["return_time_limit"]
                out["Return Time Unit"]  = STATIC["return_time_unit"]
                out["Colour"]            = color.title()
                out["Material"]          = primary_material_fabric or material_display

                out["Custom Attribute 1"]  = dept_disp
                out["Custom Attribute 2"]  = fit_display
                out["Custom Attribute 3"]  = gender
                out["Custom Attribute 4"]  = occasion.title() if occasion else ""
                out["Custom Attribute 5"]  = collar_mapped
                # Custom Attribute 6 intentionally blank per doc
                out["Custom Attribute 7"]  = sleeve_mapped
                out["Custom Attribute 8"]  = order_no
                out["Custom Attribute 9"]  = fabric_main
                out["Custom Attribute 10"] = fabric_sub
                out["Custom Attribute 11"] = fabric_sub_type
                out["Custom Attribute 12"] = hl
                out["Custom Attribute 13"] = pockets
                out["Custom Attribute 14"] = style_no
                out["Custom Attribute 20"] = fabric_no
                out["Custom Attribute 21"] = cs
                out["Custom Attribute 22"] = packed_date
                out["Custom Attribute 23"] = length_val
                out["Custom Attribute 24"] = waist_val
                out["Custom Attribute 25"] = closure
                out["Custom Attribute 26"] = leg
                out["Custom Attribute 27"] = front
                out["Custom Attribute 28"] = fabric_type
                out["Custom Attribute 29"] = rate

            output_rows.append(out)

    if not output_rows:
        raise ValueError("No valid product rows found in input file.")

    # De-duplicate warnings while preserving order
    seen: set[str] = set()
    unique_warnings: list[str] = []
    for w in warnings:
        if w not in seen:
            seen.add(w)
            unique_warnings.append(w)

    # Write output xlsx using openpyxl
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet1"
    out_ws.append(FYND_COLUMNS)
    for row_dict in output_rows:
        out_ws.append([row_dict.get(col, "") for col in FYND_COLUMNS])

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf, unique_warnings, None
