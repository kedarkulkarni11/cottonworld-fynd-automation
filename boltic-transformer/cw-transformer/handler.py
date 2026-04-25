"""
Cottonworld Logic ERP -> Fynd Platform Transformer
Single-file Boltic serverless handler.
All data inlined and uses only openpyxl + stdlib (no pandas).
"""
from __future__ import annotations

import csv
import io
import math
import re
from io import BytesIO, StringIO
from typing import Any

import openpyxl
from flask import jsonify, make_response

# ---------------------------------------------------------------------------
# Reference data (inlined — Boltic only deploys handler.py)
# ---------------------------------------------------------------------------

SLEEVE_MAP = {
    "FS": "Full Sleeves",
    "HE FS": "Full Sleeves",
    "HS": "Half Sleeves",
    "HE HS": "Half Sleeves",
    "3/4 SLEEVE WITH FOLD": "3/4 Sleeve with Fold",
    "3/4 SLV WITH BIG CUFF": "3/4 Sleeve with Big Cuff",
    "BAT SLEEVE": "Bat Sleeve",
    "ELBOW LENGTH SLV": "Elbow Length Sleeve",
    "EXTENDED SHORT SLEEVE": "Extended Short Sleeve",
    "SLEEVELESS": "Sleeveless",
    "CAP SLEEVE": "Cap Sleeve",
    "(NIL)": "",
}

COLLAR_MAP = {
    "ROUND NECK": "Rounded",
    "HOODED": "Hooded",
    "SHIRT COLLAR": "Shirt Collar",
    "BUTTON DOWN": "Button Down",
    "V NECK": "V Neck",
    "MANDARIN COLLAR": "Mandarin Collar",
    "POLO COLLAR": "Polo Collar",
    "BAND COLLAR": "Band Collar",
    "BOAT NECK": "Boat Neck",
    "SQUARE NECK": "Square Neck",
    "HIGH NECK": "High Neck",
    "(NIL)": "",
}

MATERIAL_MAP = {
    "COTTON/BAMBOO/ELASTANE": "Cotton",
    "COTTON/BAMBOO": "Cotton",
    "COTTON/BAM": "Cotton",
    "COTTON/ELASTANE": "Cotton",
    "COTTON/POLYESTER": "Cotton",
    "COTTON/LINEN": "Cotton",
    "COTTON": "Cotton",
    "LINEN/COTTON": "Linen",
    "LINEN": "Linen",
    "VISCOSE/COTTON": "Viscose",
    "VISCOSE": "Viscose",
    "POLYESTER": "Polyester",
    "SILK": "Silk",
    "WOOL": "Wool",
    "MODAL": "Modal",
    "RAYON": "Rayon",
}

SECTION_GENDER = {
    "MENS":   {"gender": "Men",   "possessive": "Men's",   "prefix": "M"},
    "LADIES": {"gender": "Women", "possessive": "Women's", "prefix": "W"},
    "BOYS":   {"gender": "Boys",  "possessive": "Boys'",   "prefix": "B"},
    "GIRLS":  {"gender": "Girls", "possessive": "Girls'",  "prefix": "G"},
    "UNISEX": {"gender": "Unisex","possessive": "Unisex",  "prefix": "U"},
}

DEPARTMENT_DISPLAY = {
    "TSHIRT": "T-shirt", "SHIRTS": "Shirt", "PANTS": "Pant",
    "SHORTS": "Shorts", "JACKET": "Jacket", "JOGGERS": "Joggers",
    "TRACK PANT": "Track Pant", "SWEAT": "Sweatshirt", "KURTA": "Kurta",
    "KURTI": "Kurti", "ETHNI KURTA": "Ethnic Kurta", "ETH KURTA": "Ethnic Kurta",
    "ETH KURTI": "Ethnic Kurti", "ETH DRESS": "Ethnic Dress",
    "ETH PALLAZO": "Ethnic Palazzo", "DRESS": "Dress", "KDRESS": "Dress",
    "SKIRT": "Skirt", "KSKIRT": "Skirt", "BLOUSE": "Blouse",
    "BOXERS": "Boxers", "BOXER": "Boxer", "CULOTTE": "Culotte",
    "JUMPER": "Jumper", "JUMPSUIT": "Jumpsuit", "KAFTAN": "Kaftan",
    "KNIT SHRUG": "Knit Shrug", "KPANTS": "Pant", "KPYJAMA SUIT": "Pyjama Suit",
    "KSHORTS": "Shorts", "KTIGHTS": "Tights", "KVEST": "Vest",
    "MASK": "Mask", "OVERLAY": "Overlay", "PAJAMA": "Pyjama",
    "PYJAMA": "Pyjama", "PYJAMA SUIT": "Pyjama Suit", "SHRUG": "Shrug",
    "SOCKS": "Socks", "WAIST COAT": "Waist Coat", "CLBAG": "Clutch Bag",
    "TOTE BAG": "Tote Bag",
}

STATIC = {
    "brand": "cottonworld",
    "category": "Others Level 3",
    "tax_rule": "Tiered Tax Rule - 5% & 18% (Eff. 22 Sep 2025) (2)",
    "country_of_origin": "India",
    "gtin_type": "EAN",
    "currency": "INR",
    "length_cm": 1,
    "width_cm": 1,
    "height_cm": 1,
    "weight_gram": 200,
    "trader_type": "Manufacturer",
    "trader_name": "Lekhraj Corp Pvt Ltd",
    "trader_address": "GALA-F, SIDHWA ESTATE, OLD BMP BUILDING, N.A. SAWANT MARG, Colaba, Mumbai City, Maharashtra, 400005",
    "return_time_limit": 30,
    "return_time_unit": "Days",
}

HSN_LOOKUP = {
    ("BOYS", "MASK"): "62171010",
    ("LADIES", "BLOUSE"): "61061000",
    ("LADIES", "BOXERS"): "61034300",
    ("LADIES", "CLBAG"): "42022240",
    ("LADIES", "CULOTTE"): "61034200",
    ("LADIES", "DRESS"): "61044200",
    ("LADIES", "ETH DRESS"): "62044911",
    ("LADIES", "ETH KURTA"): "62113900",
    ("LADIES", "ETH KURTI"): "62113900",
    ("LADIES", "ETH PALLAZO"): "62041300",
    ("LADIES", "JACKET"): "61033200",
    ("LADIES", "JUMPER"): "61044200",
    ("LADIES", "JUMPSUIT"): "62114990",
    ("LADIES", "KAFTAN"): "62114210",
    ("LADIES", "KDRESS"): "61044200",
    ("LADIES", "KNIT SHRUG"): "61033200",
    ("LADIES", "KPANTS"): "61034200",
    ("LADIES", "KPYJAMA SUIT"): "62082100",
    ("LADIES", "KSHORTS"): "61046200",
    ("LADIES", "KSKIRT"): "61045200",
    ("LADIES", "KTIGHTS"): "61034200",
    ("LADIES", "KURTA"): "62114210",
    ("LADIES", "KURTI"): "61061000",
    ("LADIES", "KVEST"): "61091000",
    ("LADIES", "MASK"): "62171000",
    ("LADIES", "OVERLAY"): "61044990",
    ("LADIES", "PAJAMA"): "62082100",
    ("LADIES", "PANTS"): "61034200",
    ("LADIES", "PYJAMA"): "62082100",
    ("LADIES", "PYJAMA SUIT"): "62082100",
    ("LADIES", "SHIRTS"): "62063000",
    ("LADIES", "SHORTS"): "61034300",
    ("LADIES", "SHRUG"): "61033200",
    ("LADIES", "SKIRT"): "61045200",
    ("LADIES", "SOCKS"): "62171010",
    ("LADIES", "TSHIRT"): "61091000",
    ("LADIES", "WAIST COAT"): "62113200",
    ("MENS", "BOXERS"): "62071100",
    ("MENS", "ETHNI KURTA"): "62113990",
    ("MENS", "JACKET"): "61033200",
    ("MENS", "JOGGERS"): "61121100",
    ("MENS", "KPANTS"): "61034200",
    ("MENS", "KSHORTS"): "61034200",
    ("MENS", "KURTA"): "62114210",
    ("MENS", "KVEST"): "62079990",
    ("MENS", "PANTS"): "61034200",
    ("MENS", "SHIRTS"): "62052000",
    ("MENS", "SHORTS"): "61034200",
    ("MENS", "SOCKS"): "62171010",
    ("MENS", "SWEAT"): "61091000",
    ("MENS", "TRACK PANT"): "61034200",
    ("MENS", "TSHIRT"): "61091000",
    ("UNISEX", "TOTE BAG"): "42022220",
}

FYND_FIXED_COLUMNS = [
    "Name", "Slug", "Item Code", "Brand", "Category", "Description",
    "Short Description", "Tax Rule Name", "HS Code", "Country of Origin",
    "Media", "Multi Size", "Gtin Type", "Gtin Value", "Seller Identifier",
    "Meta", "Size Meta", "Size", "Actual Price", "Selling Price",
    "Currency", "Length (cm)", "Width (cm)", "Height (cm)",
    "Product Dead Weight (gram)", "Size Guide", "Available", "Highlights",
    "Unlisted Product", "Variant Type", "Variant Group ID", "Variant Media",
    "Trader Type", "Trader Name", "Trader Address", "Track Inventory",
    "Teaser Tag Name", "No of Boxes", "Manufacturing Time",
    "Manufacturing Time Unit", "Return Time Limit", "Return Time Unit",
    "Product Publishing Date", "Tags", "Net Quantity Value",
    "Net Quantity Unit", "Product Bundle", "Colour", "Material",
    "Package Contents", "Quantity Factor", "Priority",
]
FYND_CUSTOM_ATTRS = [f"Custom Attribute {i}" for i in range(1, 51)]
FYND_COLUMNS = FYND_FIXED_COLUMNS + FYND_CUSTOM_ATTRS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return False


def clean_val(val: Any) -> str:
    if _is_empty(val):
        return ""
    if isinstance(val, float) and val == int(val):
        val = int(val)
    s = str(val).strip()
    if s.upper() == "(NIL)" or s.lower() == "nan":
        return ""
    return s


def strip_percentages(composition: str) -> str:
    if not composition:
        return ""
    text = re.sub(r"\d+\s*%\s*", "", composition)
    text = re.sub(r"\s+", " ", text).strip(" ,/-")
    return text.title()


def extract_primary_material(fabric_sub_type_raw: str) -> str:
    val = clean_val(fabric_sub_type_raw)
    if not val:
        return ""
    upper = val.upper()
    if upper in MATERIAL_MAP:
        return MATERIAL_MAP[upper]
    for key, mapped in MATERIAL_MAP.items():
        if key and key in upper:
            return mapped
    return val.split("/")[0].strip().title()


def map_sleeve_type(sleeve_raw: str, warnings: list) -> str:
    val = clean_val(sleeve_raw).upper()
    if not val:
        return ""
    if val in SLEEVE_MAP:
        return SLEEVE_MAP[val]
    for key, mapped in SLEEVE_MAP.items():
        if key and key in val:
            return mapped
    warnings.append(f"Unknown sleeve type '{sleeve_raw}' -- passed through as-is.")
    return val.title()


def map_collar_style(collar_raw: str, warnings: list) -> str:
    val = clean_val(collar_raw).upper()
    if not val:
        return ""
    if val in COLLAR_MAP:
        return COLLAR_MAP[val]
    warnings.append(f"Unknown neck/collar '{collar_raw}' -- passed through as-is.")
    return val.title()


def gender_info(section: str) -> dict:
    key = section.upper().strip()
    if key in SECTION_GENDER:
        return SECTION_GENDER[key]
    return {"gender": section.title(), "possessive": f"{section.title()}'s", "prefix": "X"}


def department_display(department: str) -> str:
    key = department.upper().strip()
    if key in DEPARTMENT_DISPLAY:
        return DEPARTMENT_DISPLAY[key]
    return department.title()


def build_product_name(section: str, composition1: str, fit: str,
                       department: str, color: str) -> str:
    info = gender_info(section)
    parts = [info["possessive"]]
    comp = strip_percentages(clean_val(composition1))
    if comp:
        parts.append(comp)
    fit_clean = clean_val(fit).title()
    if fit_clean:
        parts.append(fit_clean)
    dept_disp = department_display(clean_val(department))
    if dept_disp:
        parts.append(dept_disp)
    color_clean = clean_val(color).title()
    if color_clean:
        parts.append(color_clean)
    return " ".join(parts)


def build_item_code(section: str, department: str, style_no: str,
                    fabric_no: str, color: str) -> str:
    info = gender_info(section)
    prefix = info["prefix"]
    dept = clean_val(department).upper().replace(" ", "-")
    style = clean_val(style_no)
    fabric = clean_val(fabric_no)
    col = clean_val(color).upper().replace(" ", "-")
    return f"{prefix}-{dept}-{style}-{fabric}-{col}"


def lookup_hs_code(section: str, department: str, warnings: list) -> str:
    sec = clean_val(section).upper()
    dept = clean_val(department).upper()
    if not sec or not dept:
        return ""
    hs = HSN_LOOKUP.get((sec, dept))
    if hs is None:
        warnings.append(
            f"No HSN mapping for Section='{sec}', Department='{dept}'. "
            f"HS Code left blank."
        )
        return ""
    return hs


def format_size(size_raw: str) -> str:
    val = clean_val(size_raw).upper()
    size_shortforms = {
        "SMALL": "S", "MEDIUM": "M", "LARGE": "L",
        "XLARGE": "XL", "XXLARGE": "XXL", "XSMALL": "XS",
        "X-LARGE": "XL", "X-SMALL": "XS", "XX-LARGE": "XXL",
    }
    return size_shortforms.get(val, val)


def clean_barcode(raw: Any) -> str:
    val = clean_val(raw)
    if not val:
        return ""
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return val


def _normalize(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def find_col(headers: list, *candidates: str):
    norm_map = {_normalize(h): i for i, h in enumerate(headers)}
    for cand in candidates:
        n = _normalize(cand)
        if n in norm_map:
            return norm_map[n]
    return None


# ---------------------------------------------------------------------------
# Transform
# ---------------------------------------------------------------------------

def transform(input_file) -> tuple:
    warnings = []

    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    ws = wb.active
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    header_idx = None
    for i, row in enumerate(all_rows):
        for val in row:
            if str(val).strip().upper() == "OEM_BARCODE":
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is None:
        raise ValueError("Could not find 'OEM_BARCODE' header row in input file.")

    headers = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    col_oem       = find_col(headers, "OEM_BARCODE")
    col_section   = find_col(headers, "SECTION")
    col_dept      = find_col(headers, "DEPARTMENT")
    col_style_no  = find_col(headers, "STYLE NO", "STYLE_NO", "STYLENO")
    col_fabric_no = find_col(headers, "FABRIC NO.", "FABRIC_NO", "FABRICNO")
    col_color     = find_col(headers, "COLOR")
    col_size      = find_col(headers, "PACK / SIZE", "PACK_SIZE", "PACKSIZE", "PACK/SIZE")
    col_fab_main  = find_col(headers, "FABRIC MAIN DESC", "FABRIC_MAIN_DESC")
    col_fab_sub   = find_col(headers, "FABRIC SUB DESC", "FABRIC_SUB_DESC")
    col_fab_type  = find_col(headers, "FABRIC TYPE", "FABRIC_TYPE")
    col_fab_subtype = find_col(headers, "FABRIC SUB TYPE", "FABRIC_SUB_TYPE")
    col_hl        = find_col(headers, "HL")
    col_sleeve    = find_col(headers, "SLEEVE TYPE", "SLEEVE_TYPE")
    col_fit       = find_col(headers, "FIT")
    col_occasion  = find_col(headers, "OCCASION")
    col_pockets   = find_col(headers, "POCKETS")
    col_neck      = find_col(headers, "NECK-COLLAR", "NECK_COLLAR", "NECKCOLLAR")
    col_length    = find_col(headers, "LENGTH")
    col_waist     = find_col(headers, "WAIST")
    col_closure   = find_col(headers, "CLOSURE")
    col_leg       = find_col(headers, "LEG")
    col_front     = find_col(headers, "FRONT")
    col_comp1     = find_col(headers, "COMPOSITION1")
    col_packed_date = find_col(headers, "PACKED DATE", "PACKED_DATE")
    col_cs        = find_col(headers, "CS")
    col_rate      = find_col(headers, "RATE")
    col_order_no  = find_col(headers, "ORDER NO", "ORDER_NO")
    col_mrp       = find_col(headers, "MRP")

    required = {
        "OEM_BARCODE": col_oem, "SECTION": col_section, "DEPARTMENT": col_dept,
        "STYLE NO": col_style_no, "FABRIC NO.": col_fabric_no,
        "COLOR": col_color, "PACK / SIZE": col_size, "MRP": col_mrp,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"Input file missing required columns: {', '.join(missing)}")

    def get(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    valid_rows = [r for r in data_rows if not _is_empty(get(r, col_oem))]

    groups: dict = {}
    group_order = []
    for row in valid_rows:
        key = (
            clean_val(get(row, col_style_no)) + "|" +
            clean_val(get(row, col_fabric_no)) + "|" +
            clean_val(get(row, col_color))
        )
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(row)

    output_rows = []

    for key in group_order:
        group = groups[key]
        first = group[0]

        section      = clean_val(get(first, col_section))
        department   = clean_val(get(first, col_dept))
        style_no     = clean_val(get(first, col_style_no))
        fabric_no    = clean_val(get(first, col_fabric_no))
        color        = clean_val(get(first, col_color))
        fit          = clean_val(get(first, col_fit))
        occasion     = clean_val(get(first, col_occasion))
        neck_collar  = clean_val(get(first, col_neck))
        sleeve_type  = clean_val(get(first, col_sleeve))
        fabric_main  = clean_val(get(first, col_fab_main))
        fabric_sub   = clean_val(get(first, col_fab_sub))
        fabric_type  = clean_val(get(first, col_fab_type))
        fabric_sub_type = clean_val(get(first, col_fab_subtype))
        hl           = clean_val(get(first, col_hl))
        pockets      = clean_val(get(first, col_pockets))
        length_val   = clean_val(get(first, col_length))
        waist_val    = clean_val(get(first, col_waist))
        closure      = clean_val(get(first, col_closure))
        leg          = clean_val(get(first, col_leg))
        front        = clean_val(get(first, col_front))
        composition1 = clean_val(get(first, col_comp1))
        packed_date  = clean_val(get(first, col_packed_date))
        cs           = clean_val(get(first, col_cs))
        rate         = clean_val(get(first, col_rate))
        order_no     = clean_val(get(first, col_order_no))

        product_name = build_product_name(section, composition1, fit, department, color)
        item_code    = build_item_code(section, department, style_no, fabric_no, color)
        hs_code      = lookup_hs_code(section, department, warnings)
        info         = gender_info(section)
        gender       = info["gender"]
        dept_disp    = department_display(department)
        primary_material = extract_primary_material(fabric_sub_type)
        material_display = strip_percentages(composition1)
        collar_mapped = map_collar_style(neck_collar, warnings)
        sleeve_mapped = map_sleeve_type(sleeve_type, warnings)
        fit_display  = fit.title() if fit else ""

        for i, row in enumerate(group):
            is_first = (i == 0)
            oem_barcode = clean_barcode(get(row, col_oem))
            size = format_size(clean_val(get(row, col_size)))

            mrp_raw = get(row, col_mrp)
            if _is_empty(mrp_raw):
                mrp_val: Any = ""
            else:
                try:
                    mrp_val = int(float(str(mrp_raw)))
                except (ValueError, TypeError):
                    mrp_val = ""

            out = {col: "" for col in FYND_COLUMNS}
            out["Item Code"]                  = item_code
            out["Brand"]                      = STATIC["brand"]
            out["Gtin Type"]                  = STATIC["gtin_type"]
            out["Gtin Value"]                 = oem_barcode
            out["Seller Identifier"]          = oem_barcode
            out["Size"]                       = size
            out["Actual Price"]               = mrp_val
            out["Selling Price"]              = mrp_val
            out["Currency"]                   = STATIC["currency"]
            out["Length (cm)"]                = STATIC["length_cm"]
            out["Width (cm)"]                 = STATIC["width_cm"]
            out["Height (cm)"]                = STATIC["height_cm"]
            out["Product Dead Weight (gram)"] = STATIC["weight_gram"]

            if is_first:
                out["Name"]              = product_name
                out["Category"]          = STATIC["category"]
                out["Tax Rule Name"]     = STATIC["tax_rule"]
                out["HS Code"]           = hs_code
                out["Country of Origin"] = STATIC["country_of_origin"]
                out["Trader Type"]       = STATIC["trader_type"]
                out["Trader Name"]       = STATIC["trader_name"]
                out["Trader Address"]    = STATIC["trader_address"]
                out["Return Time Limit"] = STATIC["return_time_limit"]
                out["Return Time Unit"]  = STATIC["return_time_unit"]
                out["Colour"]            = color.title()
                out["Material"]          = primary_material or material_display
                out["Custom Attribute 1"]  = dept_disp
                out["Custom Attribute 2"]  = fit_display
                out["Custom Attribute 3"]  = gender
                out["Custom Attribute 4"]  = occasion.title() if occasion else ""
                out["Custom Attribute 5"]  = collar_mapped
                out["Custom Attribute 7"]  = sleeve_mapped
                out["Custom Attribute 8"]  = order_no
                out["Custom Attribute 9"]  = fabric_main
                out["Custom Attribute 10"] = fabric_sub
                out["Custom Attribute 11"] = fabric_sub_type
                out["Custom Attribute 12"] = hl
                out["Custom Attribute 13"] = pockets
                out["Custom Attribute 14"] = style_no
                out["Custom Attribute 20"] = fabric_no
                out["Custom Attribute 21"] = cs
                out["Custom Attribute 22"] = packed_date
                out["Custom Attribute 23"] = length_val
                out["Custom Attribute 24"] = waist_val
                out["Custom Attribute 25"] = closure
                out["Custom Attribute 26"] = leg
                out["Custom Attribute 27"] = front
                out["Custom Attribute 28"] = fabric_type
                out["Custom Attribute 29"] = rate

            output_rows.append(out)

    if not output_rows:
        raise ValueError("No valid product rows found in input file.")

    seen: set = set()
    unique_warnings = []
    for w in warnings:
        if w not in seen:
            seen.add(w)
            unique_warnings.append(w)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sheet1"
    out_ws.append(FYND_COLUMNS)
    for row_dict in output_rows:
        out_ws.append([row_dict.get(col, "") for col in FYND_COLUMNS])

    buf = BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf, unique_warnings


# ---------------------------------------------------------------------------
# Boltic handler
# ---------------------------------------------------------------------------

def handler(request):
    if request.method == "OPTIONS":
        res = make_response("", 204)
        res.headers["Access-Control-Allow-Origin"] = "*"
        res.headers["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        res.headers["Access-Control-Allow-Headers"] = "Content-Type"
        res.headers["Access-Control-Expose-Headers"] = "X-Warnings, Content-Disposition"
        return res

    if request.method == "GET":
        res = jsonify({"status": "ok", "version": "cw-transformer-v3"})
        res.headers["Access-Control-Allow-Origin"] = "*"
        return res

    if request.method != "POST":
        return jsonify({"detail": "Method not allowed"}), 405

    if "file" not in request.files:
        return jsonify({"detail": "No file uploaded. Send xlsx as multipart/form-data field 'file'."}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"detail": "Only .xlsx files are supported"}), 400

    try:
        contents = file.read()
        buf, warnings = transform(io.BytesIO(contents))
    except ValueError as e:
        return jsonify({"detail": str(e)}), 400
    except Exception as e:
        return jsonify({"detail": f"Transform failed: {str(e)}"}), 500

    warnings_safe = "||".join(warnings).encode("latin-1", errors="replace").decode("latin-1")

    response = make_response(buf.getvalue())
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = 'attachment; filename="fynd_catalog_output.xlsx"'
    response.headers["X-Warnings"] = warnings_safe
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Expose-Headers"] = "X-Warnings, Content-Disposition"
    return response
